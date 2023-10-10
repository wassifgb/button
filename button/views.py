import io
from django.http import HttpResponse
from urllib.parse import quote
import openpyxl
from datetime import datetime
import requests
import sys
from subprocess import run, PIPE
from django.shortcuts import render
from django.http import JsonResponse
from django.urls import reverse

from django.http import FileResponse
from django.conf import settings
from django.shortcuts import get_object_or_404
from django.conf import settings
import os

# Function to retrieve the API key from command-line arguments
def get_api_key():
    if len(sys.argv) < 2:
        print("Error: API key is missing. Please provide the API key as a command-line argument.")
        sys.exit(1)
    return sys.argv[1]

# Function to create the Excel file and return it as a downloadable file
"""def create_and_return_excel_file(api_response, excel_file_name):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value="Zone")
    worksheet.cell(row=1, column=2, value="Addresses")

    row_num = 2
    for item in api_response["result"]:
        zone = item.get("zone", "")
        addresses = item.get("addresses", [])

        addresses_str = ", ".join(addresses)

        worksheet.cell(row=row_num, column=1, value=zone)
        worksheet.cell(row=row_num, column=2, value=addresses_str)

        row_num += 1

    today = datetime.today().strftime('%d%m%Y')
    file_name_with_date = f"{excel_file_name}_{today}.xlsx"

    # Save the workbook to a bytes buffer
    output = io.BytesIO()
    workbook.save(output)

    # Create an HttpResponse with the Excel content as a downloadable file
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={quote(file_name_with_date)}'  # Use quote instead of urlquote

    print(f"Excel file '{file_name_with_date}' has been created. :)")
    
    return response"""

def create_and_return_excel_file(api_response, excel_file_name):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value="Zone")
    worksheet.cell(row=1, column=2, value="Addresses")

    row_num = 2
    for item in api_response["result"]:
        zone = item.get("zone", "")
        addresses = item.get("addresses", [])

        addresses_str = ", ".join(addresses)

        worksheet.cell(row=row_num, column=1, value=zone)
        worksheet.cell(row=row_num, column=2, value=addresses_str)

        row_num += 1

    today = datetime.today().strftime('%d%m%Y')
    file_name_with_date = f"{excel_file_name}_{today}.xlsx"

    # Modify the file path to save in MEDIA_ROOT
    file_path = os.path.join(settings.MEDIA_ROOT, file_name_with_date)

    # Save the workbook to the file path
    workbook.save(file_path)

    return file_path

# Function to fetch data and create Excel file
def fetch_data_and_create_excel(api_key, option_data, excel_file_name):
    print(f"API Key used for fetching data: {api_key}")  # Print the API key
    headers = {
        "header-api-key": api_key,
    }

    url = "https://api.prod.datapath.prismaaccess.com/getPrismaAccessIP/v2"

    response = requests.post(url, headers=headers, data=option_data)

    if response.status_code == 200:
        api_response = response.json()
        return create_and_return_excel_file(api_response, excel_file_name)
    else:
        print(f"Error: {response.status_code} - Failed to fetch data.")
        return None

# Main function
"""def main(api_key):
    print(f"API Key entered by the user: {api_key}")  # Print the API key
    print("Script start")
    # Fetch data and create Excel file for mobile users
    mobile_option_data = '''{
       "serviceType": "gp_gateway",   
       "addrType": "active",
       "location": "all"
    }'''
    response = fetch_data_and_create_excel(api_key, mobile_option_data, "Mobile_users")

    # Fetch data and create Excel file for remote networks
    remote_option_data = '''{
       "serviceType": "remote_network",
       "addrType": "active",
       "location": "all"
    }'''
    response = fetch_data_and_create_excel(api_key, remote_option_data, "Remote_networks")
    print("Script end")"""
def main(api_key):
    print(f"API Key entered by the user: {api_key}")  # Print the API key
    print("Script start")
    # Fetch data and create Excel file for mobile users
    mobile_option_data = '''{
       "serviceType": "gp_gateway",   
       "addrType": "active",
       "location": "all"
    }'''
    mobile_file_path = fetch_data_and_create_excel(api_key, mobile_option_data, "Mobile_users")

    # Fetch data and create Excel file for remote networks
    remote_option_data = '''{
       "serviceType": "remote_network",
       "addrType": "active",
       "location": "all"
    }'''
    remote_file_path = fetch_data_and_create_excel(api_key, remote_option_data, "Remote_networks")
    
    print("Script end")
    
    return mobile_file_path, remote_file_path  # Return the file paths

# Your existing code for other views
def button(request):
    return render(request, 'home.html')

def output(request):
    data = requests.get("https://www.google.com/")
    return render(request, 'home.html', {'data': data})


'''def external(request):
    if request.method == 'POST':
        inp = request.POST.get('param')
        api_key = inp  # Use the input value as the API key
        main(api_key)  # Call the main function with the API key
    return render(request, 'home.html') '''

def external(request):
    inp = request.POST.get('param')
    api_key = inp
    
    mobile_file_path, remote_file_path = main(api_key)  # Get file paths from main

    if mobile_file_path and remote_file_path:
        mobile_url = reverse('download_file', kwargs={'file_name': 'Mobile_users.xlsx'})
        remote_url = reverse('download_file', kwargs={'file_name': 'Remote_networks.xlsx'})

        return JsonResponse({
            'mobile_url': mobile_url,
            'remote_url': remote_url,
        })

    return HttpResponse("Failed to generate Excel files.")

def download_file(request, file_name):
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)
    return FileResponse(open(file_path, 'rb'), as_attachment=True)


import ipaddress
from django.http import JsonResponse

def subnet_overlap_detection(request):
    if request.method == 'POST':
        # Get user input for subnets from the POST data
        input_subnets = request.POST.get('subnets')

        if input_subnets is not None:
            # Print the input for debugging purposes
            print(f"Received input: {input_subnets}")

            # Split the input into a list of subnet strings
            input_subnets = input_subnets.split(',')

            # Convert the input subnets into IPNetwork objects and store them in a list
            subnets = []
            for subnet_str in input_subnets:
                try:
                    subnet = ipaddress.IPv4Network(subnet_str.strip())
                    subnets.append(subnet)
                except ValueError:
                    return JsonResponse({
                        'error': f"Invalid subnet format: {subnet_str}",
                    })

            # Check for overlapping subnets and identify the overlapping address range
            overlapping = False
            overlapping_pairs = []
            for i in range(len(subnets)):
                for j in range(i + 1, len(subnets)):
                    if subnets[i].overlaps(subnets[j]):
                        overlapping = True
                        overlap_start = max(subnets[i].network_address, subnets[j].network_address)
                        overlap_end = min(subnets[i].broadcast_address, subnets[j].broadcast_address)
                        overlapping_pairs.append((str(subnets[i]), str(subnets[j]), str(overlap_start), str(overlap_end)))

            if overlapping:
                overlapping_info = []
                for subnet_pair in overlapping_pairs:
                    overlapping_info.append({
                        'subnet1': subnet_pair[0],
                        'subnet2': subnet_pair[1],
                        'overlap_start': subnet_pair[2],
                        'overlap_end': subnet_pair[3],
                    })
                return JsonResponse({
                    'overlapping': True,
                    'overlapping_info': overlapping_info,
                })
            else:
                return JsonResponse({
                    'overlapping': False,
                })

    return JsonResponse({
        'error': 'Invalid request method.',
    })